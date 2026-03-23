"""
table_transfer.py

从课程表 Excel 提取排课数据，输出带时间、地点、讲师信息的 attendance Excel。

用法：
    python table_transfer.py
    python table_transfer.py --xlsx /path/to/timetable.xlsx
    python table_transfer.py --out /path/to/output.xlsx

依赖：
    pip install openpyxl
"""

import argparse
import hashlib
import json
import re
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook, Workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

# ── 默认路径 ────────────────────────────────────────────────────────
DEFAULT_XLSX = Path(__file__).parent / "副本25-26 Sem2 Timetable V1.4.xlsx"
TEMP_XLSX    = Path(__file__).parent / "temp.xlsx"   # Teaching weeks 表


# ── Week → 日期映射 ─────────────────────────────────────────────────
WEEK_DATES = {
    1:  {"Mon": "2-Mar",  "Tue": "3-Mar",  "Wed": "4-Mar",  "Thu": "5-Mar",  "Fri": "6-Mar"},
    2:  {"Mon": "9-Mar",  "Tue": "10-Mar", "Wed": "11-Mar", "Thu": "12-Mar", "Fri": "13-Mar"},
    3:  {"Mon": "16-Mar", "Tue": "17-Mar", "Wed": "18-Mar", "Thu": "19-Mar", "Fri": "20-Mar"},
    4:  {"Mon": "23-Mar", "Tue": "24-Mar", "Wed": "25-Mar", "Thu": "26-Mar", "Fri": "27-Mar"},
    5:  {"Mon": "30-Mar", "Tue": "31-Mar", "Wed": "1-Apr",  "Thu": "2-Apr",  "Fri": "3-Apr"},
    6:  {"Mon": "6-Apr",  "Tue": "7-Apr",  "Wed": "8-Apr",  "Thu": "9-Apr",  "Fri": "10-Apr"},
    7:  {"Mon": "13-Apr", "Tue": "14-Apr", "Wed": "15-Apr", "Thu": "16-Apr", "Fri": "17-Apr"},
    8:  {"Mon": "20-Apr", "Tue": "21-Apr", "Wed": "22-Apr", "Thu": "23-Apr", "Fri": "24-Apr"},
    9:  {"Mon": "27-Apr", "Tue": "28-Apr", "Wed": "29-Apr", "Thu": "30-Apr", "Fri": "1-May"},
    10: {"Mon": "4-May",  "Tue": "5-May",  "Wed": "6-May",  "Thu": "7-May",  "Fri": "8-May"},
    11: {"Mon": "11-May", "Tue": "12-May", "Wed": "13-May", "Thu": "14-May", "Fri": "15-May"},
    12: {"Mon": "18-May", "Tue": "19-May", "Wed": "20-May", "Thu": "21-May", "Fri": "22-May"},
    13: {"Mon": "25-May", "Tue": "26-May", "Wed": "27-May", "Thu": "28-May", "Fri": "29-May"},
    14: {"Mon": "1-Jun",  "Tue": "2-Jun",  "Wed": "3-Jun",  "Thu": "4-Jun",  "Fri": "5-Jun"},
    15: {"Mon": "8-Jun",  "Tue": "9-Jun",  "Wed": "10-Jun", "Thu": "11-Jun", "Fri": "12-Jun"},
}

_PROGRAMME_CLASS_MAP = {
    "Tele-G1": "1-5",
    "Tele-G2": "6-10",
    "IoT1":    "11-13",
    "IoT2":    "14-16",
    "EIE":     "17-20",
    "IST":     "21-24",
    "IoT":     "11-16",
}

DAY_ABBR_TO_FULL = {
    "Mon": "Monday", "Tue": "Tuesday", "Wed": "Wednesday",
    "Thu": "Thursday", "Fri": "Friday", "Sat": "Saturday", "Sun": "Sunday",
}


# ── 通用工具 ────────────────────────────────────────────────────────

def parse_cell(cell_str: str) -> tuple[int, int]:
    """'A4' → (row=4, col=1)"""
    col_str = re.findall(r'[A-Za-z]+', cell_str)[0]
    row_str = re.findall(r'\d+', cell_str)[0]
    return int(row_str), column_index_from_string(col_str)


def parse_region(region_str: str) -> tuple[str, str]:
    """'A2:G7' → ('A2', 'G7')"""
    tl, br = region_str.replace(" ", "").split(":")
    return tl, br


def normalize_programme(raw: str) -> str:
    s = str(raw).strip().replace(" ", "")
    s = s.replace("InternetofThings", "IoT")
    s = s.replace("TeleG1", "Tele-G1").replace("TeleG2", "Tele-G2")
    s = s.replace("TelecomG1", "Tele-G1").replace("TelecomG2", "Tele-G2")
    s = s.replace("Telecom_Y3_G1", "Tele-G1").replace("Telecom_Y3_G2", "Tele-G2")
    s = s.replace("Y3_", "").replace("Telecom_", "Tele-")
    s = s.replace("IoT_G2", "IoT2").replace("IoT_G1", "IoT1")
    return s


def get_weekday_en(date_str: str) -> str:
    """'2-Mar' → 'Monday'"""
    dt = datetime.strptime(f"2026-{date_str}", "%Y-%d-%b")
    return dt.strftime("%A")


def class_range_to_student_count(class_range: str) -> int:
    text = str(class_range).strip().replace(" ", "").replace("～", "~").replace("~", "-")
    nums = [int(x) for x in re.findall(r"\d+", text)]
    if not nums:
        return 0
    start, end = (nums[0], nums[0]) if len(nums) == 1 else (nums[0], nums[1])
    if start > end:
        start, end = end, start

    def class_size(n: int) -> int:
        if 1 <= n <= 16:
            return 30
        if 17 <= n <= 24:
            return 25
        raise ValueError(f"班级号超出范围: {n}")

    return sum(class_size(i) for i in range(start, end + 1))


def find_matched_fill_cells(ws, target_cell: str) -> list[dict]:
    """找与 target_cell 填充色相同的单元格（仅在其右侧 1-5 列、下方行）。"""
    def fill_key(cell):
        f = cell.fill
        fg, bg = f.fgColor, f.bgColor
        return (
            f.patternType,
            fg.type, fg.rgb, fg.theme, fg.indexed, fg.tint,
            bg.type, bg.rgb, bg.theme, bg.indexed, bg.tint,
        )

    target_col_str, target_row = coordinate_from_string(target_cell)
    target_col = column_index_from_string(target_col_str)
    target_key = fill_key(ws[target_cell])

    min_col = target_col + 1
    max_col = min(target_col + 6, ws.max_column)

    matched = []
    for row in ws.iter_rows(min_row=target_row + 1, min_col=min_col, max_col=max_col):
        for cell in row:
            try:
                if fill_key(cell) == target_key:
                    matched.append({"cell": cell.coordinate, "room": cell.value})
            except Exception:
                continue

    def coord_key(coord):
        col_l, row_n = coordinate_from_string(coord)
        return column_index_from_string(col_l), row_n

    return sorted(matched, key=lambda x: coord_key(x["cell"]))


# ── Step 1: 提取班级信息（class regions） ───────────────────────────

CLASS_REGIONS = [
    "A2:G7", "J2:P7", "S2:Y7", "AA2:AG7", "AI2:AO7", "AQ2:AW7",
]


def extract_class_info(ws) -> dict:
    """从 Year2/Year3 sheet 的班级区域提取模块信息。"""
    result = {}
    for region in CLASS_REGIONS:
        tl, br = parse_region(region)
        start_row, start_col = parse_cell(tl)
        end_row, _           = parse_cell(br)

        programme = normalize_programme(
            ws.cell(row=start_row, column=start_col).value or "Unknown"
        )

        for row in range(4, end_row + 1):
            code      = ws.cell(row=row, column=start_col).value
            module    = ws.cell(row=row, column=start_col + 1).value
            class_val = ws.cell(row=row, column=start_col + 5).value
            cell_idx  = ws.cell(row=row, column=start_col).coordinate

            if code:
                key = f"{programme}_{code}"
                result[key] = {
                    "Code": code,
                    "cell_index": cell_idx,
                    "Module Name": module,
                    "Class": class_val,
                    "Programme": programme,
                }
    return result


# ── Step 2: 提取单元格→日期/时间/周 映射 ────────────────────────────

YEAR2_CELL_REGIONS = [
    "B10:G20",   "B22:G32",   "B34:G44",   "B46:G56",   "B58:G68",
    "B70:G80",   "B82:G92",   "B94:G104",  "B106:G116", "B118:G128",
    "B130:G140", "B142:G152",
    "K10:P20",   "K22:P32",   "K34:P44",   "K46:P56",   "K58:P68",
    "K70:P80",   "K82:P92",   "K94:P104",  "K106:P116", "K118:P128",
    "K130:P140", "K142:P152",
    "T10:Y20",   "T22:Y32",   "T34:Y44",   "T46:Y56",   "T58:Y68",
    "T70:Y80",   "T82:Y92",   "T94:Y104",  "T106:Y116", "T118:Y128",
    "T130:Y140", "T142:Y152",
    "AB10:AG20", "AB22:AG32", "AB34:AG44", "AB46:AG56", "AB58:AG68",
    "AB70:AG80", "AB82:AG92", "AB94:AG104","AB106:AG116","AB118:AG128",
    "AB130:AG140","AB142:AG152",
    "AJ10:AO20", "AJ22:AO32", "AJ34:AO44", "AJ46:AO56", "AJ58:AO68",
    "AJ70:AO80", "AJ82:AO92", "AJ94:AO104","AJ106:AO116","AJ118:AO128",
    "AJ130:AO140","AJ142:AO152",
    "AR10:AW20", "AR22:AW32", "AR34:AW44", "AR46:AW56", "AR58:AW68",
    "AR70:AW80", "AR82:AW92", "AR94:AW104","AR106:AW116","AR118:AW128",
    "AR130:AW140","AR142:AW152",
]

# Year3 与 Year2 区域相同
YEAR3_CELL_REGIONS = YEAR2_CELL_REGIONS


def extract_cell_map(ws, regions: list[str]) -> dict:
    """构建 cell_name → {Date, Time, Week} 映射。"""
    result = {}
    for region in regions:
        tl, br = parse_region(region)
        start_row, start_col = parse_cell(tl)
        end_row, end_col     = parse_cell(br)

        # 日期行
        date_map = {}
        for col in range(start_col + 1, end_col + 1):
            val = ws.cell(row=start_row, column=col).value
            if isinstance(val, datetime):
                date_map[col] = f"{val.day}-{val.strftime('%b')}"
            else:
                date_map[col] = str(val).strip() if val else ""

        week_val = str(
            ws.cell(row=start_row, column=start_col - 1).value or "1"
        ).strip().replace("Week", "").replace("week", "").replace(" ", "")

        for row in range(start_row + 1, end_row + 1):
            time_val = str(ws.cell(row=row, column=start_col).value or "").strip()
            for col in range(start_col + 1, end_col + 1):
                cell_name = ws.cell(row=row, column=col).coordinate
                result[cell_name] = {
                    "Date": date_map[col],
                    "Time": time_val,
                    "Week": week_val,
                }
    return result


# ── Step 3: 讲师信息（temp.xlsx） ───────────────────────────────────

def load_lecturer_info(temp_path: Path) -> dict:
    wb = load_workbook(str(temp_path), data_only=True)
    ws = wb.active

    headers = {}
    for col_idx, cell in enumerate(ws[1], 1):
        headers[cell.value] = col_idx

    col_stream = headers["Stream"]
    col_code   = headers["Code"]
    col_first  = headers["First-half"]
    col_second = headers["Second-half"]
    week_cols  = [headers[w] for w in range(1, 16)]

    result = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        stream = str(row[col_stream - 1]) if row[col_stream - 1] is not None else ""
        code   = str(row[col_code - 1])   if row[col_code - 1]   is not None else ""
        stream = normalize_programme(stream)
        key = f"{stream}_{code}"

        week_data = {}
        for i, w_col in enumerate(week_cols):
            val = row[w_col - 1]
            val = str(val) if val is not None else ""
            if val.lower() == "none":
                val = ""
            week_data[f"Week_{i + 1}"] = val

        result[key] = {
            "first_half":   str(row[col_first  - 1]) if row[col_first  - 1] else "",
            "second_half":  str(row[col_second - 1]) if row[col_second - 1] else "",
            "week_1_to_15": week_data,
        }

    wb.close()
    return result


def lookup_lecturer(lecturer_info: dict, programme: str, module_code: str, week: str | int) -> str:
    week_key = f"Week_{week}"
    candidates = [f"{programme}_{module_code}", f"{programme}{module_code}"]
    if programme in {"IoT1", "IoT2"}:
        candidates.append(f"IoT_{module_code}")
    candidates.append(f"all_{module_code}")

    for key in candidates:
        info = lecturer_info.get(key)
        if not info:
            continue
        lecturer = info.get("week_1_to_15", {}).get(week_key, "")
        if lecturer:
            return lecturer
    return ""


# ── Step 4: 构建 Year2/Year3 记录 ───────────────────────────────────

def build_year_records(excel_file: str, sheet_name: str, year: str,
                       cell_regions: list[str], lecturer_info: dict) -> list[dict]:
    wb_styles = load_workbook(excel_file)               # 读样式（填充色）
    wb_data   = load_workbook(excel_file, data_only=True)  # 读数值（日期公式）
    ws_styles = wb_styles[sheet_name]
    ws_data   = wb_data[sheet_name]

    class_info = extract_class_info(ws_styles)
    cell_map   = extract_cell_map(ws_data, cell_regions)

    records = []
    for value in class_info.values():
        matched = find_matched_fill_cells(ws_styles, value["cell_index"])
        for item in matched:
            if item["cell"] not in cell_map:
                continue
            cm = cell_map[item["cell"]]
            programme  = value["Programme"]
            module_code = value["Code"]
            week = cm["Week"]
            lecturer = lookup_lecturer(lecturer_info, programme, module_code, week)

            records.append({
                "Index No.": "",
                "Week":      week,
                "Day":       get_weekday_en(cm["Date"]),
                "Date":      cm["Date"],
                "Time":      cm["Time"],
                "Room":      item["room"],
                "Module Code":   module_code,
                "Module Name":   value["Module Name"],
                "Lecturer":      lecturer,
                "Year":          year,
                "Programme":     programme,
                "Class":         str(value["Class"]),
                "Total student num":         class_range_to_student_count(str(value["Class"])),
                "Student number in classroom": 0,
                "Percent": 0,
                "By": "",
                "Remark": "",
            })
    return records


# ── Step 5: PDP ─────────────────────────────────────────────────────

PDP_CLASSES = {
    "PDP2_FL": {"Code": "EBC4002", "cell_index": "B65", "Lecturer": "FL", "Module Name": "PDP2"},
    "PDP2_NY": {"Code": "EBC4002", "cell_index": "C65", "Lecturer": "NY", "Module Name": "PDP2"},
    "PDP3_BC": {"Code": "EBC5002", "cell_index": "B115", "Lecturer": "BC", "Module Name": "PDP3"},
    "PDP3_MM": {"Code": "EBC5002", "cell_index": "C115", "Lecturer": "MM", "Module Name": "PDP3"},
}

PDP_CELL_REGIONS = [
    "B66:G76",  "B78:G88",  "B90:G100",  "B102:G112",
    "B116:G126","B128:G138","B140:G150","B152:G162",
]


def build_pdp_records(excel_file: str) -> list[dict]:
    wb_styles = load_workbook(excel_file)
    wb_data   = load_workbook(excel_file, data_only=True)
    ws_styles = wb_styles["PDP"]
    ws_data   = wb_data["PDP"]

    # PDP 的 cell_map：Date 列直接是星期缩写（Mon/Tue...），不是日期
    cell_map = {}
    for region in PDP_CELL_REGIONS:
        tl, br = parse_region(region)
        start_row, start_col = parse_cell(tl)
        end_row, end_col     = parse_cell(br)

        date_map = {}
        for col in range(start_col + 1, end_col + 1):
            val = ws_data.cell(row=start_row, column=col).value
            date_map[col] = str(val).strip() if val else ""

        week_val = str(
            ws_data.cell(row=start_row, column=start_col - 1).value or "1"
        ).strip().replace("Week", "").replace("week", "").replace(" ", "")

        for row in range(start_row + 1, end_row + 1):
            time_val = str(ws_data.cell(row=row, column=start_col).value or "").strip()
            for col in range(start_col + 1, end_col + 1):
                cell_name = ws_data.cell(row=row, column=col).coordinate
                cell_map[cell_name] = {
                    "Day": date_map[col],   # e.g. "Mon"
                    "Time": time_val,
                    "Week": week_val,
                }

    records = []
    for pdp in PDP_CLASSES.values():
        matched = find_matched_fill_cells(ws_styles, pdp["cell_index"])
        for item in matched:
            if item["cell"] not in cell_map:
                continue
            cm = cell_map[item["cell"]]
            week = int(cm["Week"])
            day_abbr = cm["Day"]
            date = WEEK_DATES[week][day_abbr]

            class_match = re.search(r'class(\d{2})', str(item["room"]))
            room_match  = re.search(r'3-\d{3}', str(item["room"]))
            class_no    = class_match.group(1) if class_match else ""
            room_info   = room_match.group()   if room_match  else ""

            records.append({
                "Index No.": "",
                "Week":      str(week),
                "Day":       DAY_ABBR_TO_FULL[day_abbr],
                "Date":      date,
                "Time":      cm["Time"],
                "Room":      room_info,
                "Module Code":   pdp["Code"],
                "Module Name":   pdp["Module Name"],
                "Lecturer":      pdp["Lecturer"],
                "Year":          "2" if "2" in pdp["Module Name"] else "3",
                "Programme":     "",
                "Class":         class_no,
                "Total student num":          class_range_to_student_count(str(int(class_no))) if class_no else 0,
                "Student number in classroom": 0,
                "Percent": 0,
                "By": "",
                "Remark": "",
            })
    return records


# ── Step 6: Tutorial ─────────────────────────────────────────────────

def build_tutorial_records(excel_file: str, lecturer_info: dict) -> list[dict]:
    wb = load_workbook(excel_file, data_only=True)
    ws = wb["OH-Tut"]

    data_iter = ws.values
    headers = next(data_iter)
    rows = [dict(zip(headers, row)) for row in data_iter]
    rows = rows[1:]  # 跳过第一行（通常是空行或说明）

    records = []
    for item in rows:
        bupt_week = item.get("BUPT Week:")
        if bupt_week is None:
            continue
        weeks = [str(bupt_week)] if isinstance(bupt_week, int) else str(bupt_week).split(",")

        tut_time = str(item.get("Tutorial time") or "").strip()
        m = re.match(r"([A-Za-z]+)\.*(\d{1,2}[:.]\d{2}-\d{1,2}[:.]\d{2})", tut_time)
        if not m:
            continue
        day_raw = m.group(1).replace("Thurs", "Thu").replace("Thur", "Thu")
        time    = m.group(2).replace(".", ":")

        programme  = normalize_programme(str(item.get("Stream") or ""))
        module_code = str(item.get("Code") or "").strip()
        room_info   = str(item.get("Tutorial room") or "").strip()
        class_info  = _PROGRAMME_CLASS_MAP.get(programme, "")

        for week in weeks:
            week = week.strip()
            if not week:
                continue
            date = WEEK_DATES[int(week)][day_raw]
            lecturer = lookup_lecturer(lecturer_info, programme, module_code, week)

            records.append({
                "Index No.": "",
                "Week":      week,
                "Day":       DAY_ABBR_TO_FULL[day_raw],
                "Date":      date,
                "Time":      time,
                "Room":      room_info,
                "Module Code":   module_code,
                "Module Name":   str(item.get("Module Name") or "").strip(),
                "Lecturer":      lecturer,
                "Year":          str(item.get("Year") or "").replace("Year ", "").strip(),
                "Programme":     programme,
                "Class":         class_info,
                "Total student num":          class_range_to_student_count(class_info) if class_info else 0,
                "Student number in classroom": 0,
                "Percent": 0,
                "By": "",
                "Remark": "Tutorial",
            })
    return records


# ── Step 7: 排序 ─────────────────────────────────────────────────────

def _to_int(v, default=10**9):
    try:
        return int(str(v).strip())
    except Exception:
        return default


def _day_order(day: str) -> int:
    return {"Monday":1,"Tuesday":2,"Wednesday":3,"Thursday":4,"Friday":5}.get(str(day).strip(), 99)


def _time_key(t: str) -> tuple:
    m = re.match(r"^(\d{1,2}):(\d{2})", str(t or "").strip())
    return (int(m.group(1)), int(m.group(2))) if m else (99, 99)


def _room_key(room: str) -> tuple:
    text = str(room or "").strip()
    prefix = 0 if text.startswith("3-") else (1 if text.startswith("4-") else 2)
    parts = text.split("-", 1)
    nums = re.findall(r"\d+", parts[1]) if len(parts) == 2 else []
    return (prefix, -int(nums[0]) if nums else 0, text)


def sort_records(records: list[dict]) -> list[dict]:
    return sorted(records, key=lambda x: (
        _to_int(x.get("Week")),
        _day_order(x.get("Day")),
        _time_key(x.get("Time")),
        _room_key(x.get("Room")),
    ))


# ── Step 8: 导出 Excel ───────────────────────────────────────────────

def export_to_excel(records: list[dict], output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    if not records:
        wb.save(output_path)
        return
    headers = list(records[0].keys())
    ws.append(headers)
    for row in records:
        ws.append([row.get(h, "") for h in headers])
    wb.save(output_path)
    print(f"✅ 已导出：{output_path}  ({len(records)} 条记录)")


# ── 主流程 ───────────────────────────────────────────────────────────

def main(excel_file: str, out_path: str | None):
    print(f"读取：{excel_file}")

    lecturer_info = load_lecturer_info(TEMP_XLSX)
    print(f"  讲师信息：{len(lecturer_info)} 条")

    print("  提取 Year2 ...")
    year2 = build_year_records(excel_file, "Year2", "2", YEAR2_CELL_REGIONS, lecturer_info)
    print(f"    {len(year2)} 条")

    print("  提取 Year3 ...")
    year3 = build_year_records(excel_file, "Year3", "3", YEAR3_CELL_REGIONS, lecturer_info)
    print(f"    {len(year3)} 条")

    print("  提取 PDP ...")
    pdp = build_pdp_records(excel_file)
    print(f"    {len(pdp)} 条")

    print("  提取 Tutorial ...")
    tutorial = build_tutorial_records(excel_file, lecturer_info)
    print(f"    {len(tutorial)} 条")

    total = year2 + year3 + pdp + tutorial
    total = sort_records(total)

    # 后处理
    for item in total:
        if item["Class"] == "11-16":
            item["Programme"] = "IoT"
        item["Week"]  = str(item["Week"])
        item["Year"]  = str(item["Year"])
        item["Class"] = str(item["Class"]).replace("~", "-")

    # 补序号 + ID
    for i, item in enumerate(total):
        item["Index No."] = i + 1
        hash_str = f"{item['Date']}{item['Time']}{item['Room']}"
        item["ID"] = hashlib.md5(hash_str.encode()).hexdigest()

    print(f"  合计：{len(total)} 条")

    if out_path is None:
        timestamp = datetime.now().strftime("%Y-%m-%d-%H%M")
        out_path = str(Path(__file__).parent / f"{timestamp}-export.xlsx")

    export_to_excel(total, out_path)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="从课程表 Excel 生成 attendance 数据")
    parser.add_argument("--xlsx", default=str(DEFAULT_XLSX), help="课程表 Excel 路径")
    parser.add_argument("--out",  default=None,              help="输出 Excel 路径")
    args = parser.parse_args()
    main(args.xlsx, args.out)
